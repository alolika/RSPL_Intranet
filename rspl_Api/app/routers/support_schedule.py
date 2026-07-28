"""Executive Schedule (Section_Support/GSchedule.aspx replacement).

The source page pulled the "Schedule" tab of a Google Sheet via Google's old
ClientLogin API, retired years ago — genuinely unreachable, not a code bug
(see g-schedule.ts's docstring).

Primary data path: Power Automate. A flow (built by the user, not this repo)
watches Amol's Outlook mailbox for the Schedule.xlsx attachment and POSTs the
raw attachment bytes as the request body straight to
POST /executive-schedule/import, authenticated with a shared secret
(X-Import-Key header, see settings.schedule_import_api_key). Raw bytes rather
than a multipart file upload deliberately, since Power Automate's HTTP action
can pass an attachment's contentBytes directly as the body with no manual
multipart-boundary construction — that's a common Power Automate pain point
this sidesteps entirely.
The parsed {columns, rows} snapshot is cached to a local JSON file and served
back to the frontend by GET /executive-schedule on every page load — this
mirrors how source column data is genuinely a full-refresh "current status"
snapshot each time (whoever's in the office/on leave/on holiday *right now*),
not an accumulating table, so each import simply replaces the previous
snapshot wholesale rather than needing per-row insert/update/duplicate logic.

Fallback data path: a Microsoft Graph app-only pull directly from Amol's
OneDrive, kept from an earlier attempt at automating this without Power
Automate. Reading a *personal* OneDrive (as opposed to a Team Site library)
via pure app-only Graph permissions needs two extra tenant-side grants beyond
the usual Entra ID consent flow — the `Sites.Selected` application permission
plus a one-time PnP PowerShell `Grant-PnPAzureADAppSitePermission` scoping it
to this one site — which were still unresolved (token had no roles at all)
when the Power Automate approach was adopted instead. Left in place as a
fallback only used when no Power Automate import has landed yet; if it starts
mattering again, check those two grants first.

No column names are hardcoded anywhere in this file: whatever's in row 1
becomes the columns, matching GSchedule's already-generic {columns, rows}
contract (g-schedule.ts's cellColor() treats columns 0-2 as plain labels and
3+ as O/L/H-coded status cells purely by position, not by name).
"""

import base64
import json
import logging
import time
from io import BytesIO
from pathlib import Path

import requests
from fastapi import APIRouter, Header, HTTPException, Request
from openpyxl import load_workbook
from pydantic import BaseModel

from app.config import settings

router = APIRouter(prefix="/support", tags=["support"])

logger = logging.getLogger(__name__)

GRAPH_TOKEN_URL = "https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token"
GRAPH_SHARE_CONTENT_URL = "https://graph.microsoft.com/v1.0/shares/{share_id}/driveItem/content"

# Same C:\Retailware convention as support.py's TT attachment storage — this
# backend's own process runs on that same class of Windows machine.
CACHE_FILE = Path(r"C:\Retailware\ExecutiveSchedule\schedule_data.json")

# Simple in-memory cache — app-only tokens are valid ~1hr and the Graph
# fallback path is low-traffic; no need to hit the token endpoint every load.
_token_cache: dict[str, object] = {"access_token": None, "expires_at": 0.0}


class ExecutiveScheduleResponse(BaseModel):
    available: bool
    columns: list[str]
    rows: list[dict[str, str]]


class ImportResponse(BaseModel):
    success: bool
    rows_imported: int


def _parse_schedule_workbook(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    header_row = next(rows_iter, None)
    if not header_row:
        return [], []

    # Trailing empty header cells happen when the sheet's used-range is wider
    # than the actual data (common in hand-maintained sheets) — drop them
    # rather than emitting blank-named columns.
    columns: list[str] = []
    for cell in header_row:
        if cell is None or str(cell).strip() == "":
            break
        columns.append(str(cell).strip())

    rows: list[dict[str, str]] = []
    for raw_row in rows_iter:
        if raw_row is None or all(v is None or str(v).strip() == "" for v in raw_row[: len(columns)]):
            continue  # skip fully-blank rows (trailing rows in the used range, etc.)
        row_dict: dict[str, str] = {}
        for i, col in enumerate(columns):
            value = raw_row[i] if i < len(raw_row) else None
            row_dict[col] = "" if value is None else str(value).strip()
        rows.append(row_dict)

    return columns, rows


def _save_cache(columns: list[str], rows: list[dict[str, str]]) -> None:
    CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
    CACHE_FILE.write_text(json.dumps({"columns": columns, "rows": rows}), encoding="utf-8")


def _load_cache() -> ExecutiveScheduleResponse | None:
    if not CACHE_FILE.exists():
        return None
    try:
        data = json.loads(CACHE_FILE.read_text(encoding="utf-8"))
        return ExecutiveScheduleResponse(available=bool(data.get("columns")), columns=data.get("columns", []), rows=data.get("rows", []))
    except Exception:
        logger.exception("Executive Schedule: cache file exists but couldn't be read")
        return None


def _encode_share_id(share_url: str) -> str:
    """Encodes a sharing URL into the 'u!...' shareId format Graph's
    /shares endpoint expects. See Microsoft's "Access shared DriveItems"
    Graph docs — base64, then made URL-safe, padding stripped."""
    b64 = base64.b64encode(share_url.encode("utf-8")).decode("utf-8")
    return "u!" + b64.replace("=", "").replace("/", "_").replace("+", "-")


def _get_graph_token() -> str | None:
    if not (settings.graph_client_id and settings.graph_tenant_id and settings.graph_client_secret):
        return None

    now = time.time()
    if _token_cache["access_token"] and now < _token_cache["expires_at"]:
        return _token_cache["access_token"]  # type: ignore[return-value]

    resp = requests.post(
        GRAPH_TOKEN_URL.format(tenant=settings.graph_tenant_id),
        data={
            "client_id": settings.graph_client_id,
            "client_secret": settings.graph_client_secret,
            "scope": "https://graph.microsoft.com/.default",
            "grant_type": "client_credentials",
        },
        timeout=15,
    )
    resp.raise_for_status()
    body = resp.json()
    token = body["access_token"]
    _token_cache["access_token"] = token
    _token_cache["expires_at"] = now + body.get("expires_in", 3600) - 60
    return token


def _fetch_from_graph() -> ExecutiveScheduleResponse | None:
    if not settings.graph_schedule_share_url:
        return None
    try:
        token = _get_graph_token()
        if not token:
            return None

        share_id = _encode_share_id(settings.graph_schedule_share_url)
        resp = requests.get(
            GRAPH_SHARE_CONTENT_URL.format(share_id=share_id),
            headers={"Authorization": f"Bearer {token}"},
            timeout=30,
        )
        resp.raise_for_status()

        columns, rows = _parse_schedule_workbook(resp.content)
        return ExecutiveScheduleResponse(available=bool(columns), columns=columns, rows=rows)
    except Exception:
        # Same honest-degrade philosophy as the rest of this app's
        # best-effort integrations (SupportCoOrdinatorDashboard's per-tile
        # try/except, MyCaller's call-location enrichment, etc.) — a broken
        # external source shows "not available", not a 500.
        logger.exception("Executive Schedule: Graph fallback failed to fetch/parse Schedule.xlsx")
        return None


@router.get("/executive-schedule", response_model=ExecutiveScheduleResponse)
def get_executive_schedule() -> ExecutiveScheduleResponse:
    cached = _load_cache()
    if cached is not None:
        return cached

    from_graph = _fetch_from_graph()
    if from_graph is not None:
        return from_graph

    return ExecutiveScheduleResponse(available=False, columns=[], rows=[])


@router.post("/executive-schedule/import", response_model=ImportResponse)
async def import_executive_schedule(request: Request, x_import_key: str = Header(...)) -> ImportResponse:
    if not settings.schedule_import_api_key or x_import_key != settings.schedule_import_api_key:
        raise HTTPException(status_code=401, detail="Invalid import key")

    # Raw bytes, not a multipart upload — see module docstring for why.
    content = await request.body()
    columns, rows = _parse_schedule_workbook(content)
    if not columns:
        raise HTTPException(status_code=400, detail="Could not read any columns from the uploaded file")

    _save_cache(columns, rows)
    return ImportResponse(success=True, rows_imported=len(rows))
